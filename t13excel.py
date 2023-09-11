import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb=xl.load_workbook("filename")
    sheet=wb["Sheet 1"]


    for row in range(1,sheet.max_rox +1):
       cell+sheet.cell(row,3)
       corrected_price=cell.value*0.9
       corrected_pricw_cell=sheet.cell(row4)
       corrected_pricw_cell.value=corrected_price

    values= Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart=BarChart
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)
